import json, os
from sqlalchemy.orm import Session
from openpyxl import load_workbook


from models import ModelsOriginClass
from utils import BASE_DIR


def insert_origin_class_from_xlsx_old(db: Session):
    # 导入一个工作簿
    wb = load_workbook(f"{BASE_DIR}/course_origin.xlsx")
    # 获取当前活动的工作表
    ws = wb.active

    max_row = ws.max_row

    for rows in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
        origin_data = list(rows)
        # print(origin_data)
        if (
            not db.query(ModelsOriginClass)
            .filter(ModelsOriginClass.teacher_name == origin_data[0])
            .filter(ModelsOriginClass.course_name == origin_data[1])
            .filter(ModelsOriginClass.class_name == origin_data[2])
            .first()
        ):
            OriginClass = ModelsOriginClass(
                teacher_name=origin_data[0],
                course_name=origin_data[1],
                class_name=origin_data[2],
                population=origin_data[3],
                software_name=origin_data[4],
                cycle=origin_data[5],
                teacher_room=origin_data[6],
                week_lesson_times=origin_data[7],
            )
            db.add(OriginClass)
            db.commit()
    os.remove(f"{BASE_DIR}/course_origin.xlsx")


# def insert_origin_courses_from_json(db: Session):
#     with open(f"{BASE_DIR}/course_origin.json", "r", encoding="utf-8") as f:
#         courses = json.load(f)["data"]

#     for course in courses:
#         # 检查是否存在相同的记录
#         existing_course = (
#             db.query(ModelsOriginClass)
#             .filter(
#                 ModelsOriginClass.teacherName == course["teacherName"],
#                 ModelsOriginClass.courseName == course["courseName"],
#                 ModelsOriginClass.className == course["className"],
#             )
#             .first()
#         )

#         if not existing_course:
#             new_course = ModelsOriginClass(**course)
#             db.add(new_course)

#     db.commit()


# def insert_main_courses_from_json(db: Session):
#     with open(f"{BASE_DIR}/course_main.json", "r", encoding="utf-8") as f:
#         courses = json.load(f)["data"]

#     for course in courses:
#         # 检查是否存在相同的记录
#         existing_course = (
#             db.query(ModelsMainClass)
#             .filter(
#                 ModelsMainClass.teacherName == course["teacherName"],
#                 ModelsMainClass.courseName == course["courseName"],
#                 ModelsMainClass.className == course["className"],
#             )
#             .first()
#         )

#         if not existing_course:
#             new_course = ModelsMainClass(**course)
#             db.add(new_course)

#     db.commit()


# def insert_floor_from_json(db: Session):
#     with open(f"{BASE_DIR}/floor.json", "r", encoding="utf-8") as f:
#         floors = json.load(f)["floors"]
#     for floor in floors:
#         for i in floors[f"{floor}"]:
#             # 检查是否存在相同的记录
#             existing_floor = (
#                 db.query(Floors)
#                 .filter(
#                     Floors.classRoomName == i,
#                 )
#                 .first()
#             )
#             if not existing_floor:
#                 new_floor = Floors(classRoomName=i)
#                 db.add(new_floor)

#     db.commit()
