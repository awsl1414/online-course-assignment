import json, os
from sqlalchemy.orm import Session
from openpyxl import load_workbook

# 解决循环导入
import models


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def insert_origin_courses_from_json(db: Session):
    with open(f"{BASE_DIR}/course_origin.json", "r", encoding="utf-8") as f:
        courses = json.load(f)["data"]

    for course in courses:
        # 检查是否存在相同的记录
        existing_course = (
            db.query(models.OriginClass)
            .filter(
                models.OriginClass.teacherName == course["teacherName"],
                models.OriginClass.courseName == course["courseName"],
                models.OriginClass.className == course["className"],
            )
            .first()
        )

        if not existing_course:
            new_course = models.OriginClass(**course)
            db.add(new_course)

    db.commit()


def insert_main_courses_from_json(db: Session):
    with open(f"{BASE_DIR}/course_main.json", "r", encoding="utf-8") as f:
        courses = json.load(f)["data"]

    for course in courses:
        # 检查是否存在相同的记录
        existing_course = (
            db.query(models.MainClass)
            .filter(
                models.MainClass.teacherName == course["teacherName"],
                models.MainClass.courseName == course["courseName"],
                models.MainClass.className == course["className"],
            )
            .first()
        )

        if not existing_course:
            new_course = models.MainClass(**course)
            db.add(new_course)

    db.commit()


def insert_origin_courses_from_xlsx(db: Session):
    # 导入一个工作簿
    wb = load_workbook("backend/course_origin.xlsx")
    # 获取当前活动的工作表
    ws = wb.active

    max_row = ws.max_row

    for rows in ws.iter_rows(min_row=2, max_row=max_row):
        origin_data = []
        for row in rows:
            origin_data.append(row.value)
        print(origin_data)
        if (
            not db.query(models.OriginClass)
            .filter(models.OriginClass.teacherName == origin_data[0])
            .filter(models.OriginClass.courseName == origin_data[1])
            .filter(models.OriginClass.className == origin_data[2])
            .first()
        ):
            OriginClass = models.OriginClass(
                teacherName=origin_data[0],
                courseName=origin_data[1],
                className=origin_data[2],
                population=origin_data[3],
                software=origin_data[4],
                # week=origin_data[5],
                teacherRoom=origin_data[6],
                cycle=origin_data[7],
            )
            db.add(OriginClass)
            db.commit()
    os.remove("backend/course_origin.xlsx")
