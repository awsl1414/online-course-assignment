import io
from typing import Type
from sqlalchemy.orm import Session, DeclarativeMeta
from openpyxl import load_workbook


def insert_test_from_xlsx(db: Session, file_content: io.BytesIO):
    wb = load_workbook(filename=file_content)
    ws = wb.active

    # 循环"机房"列，打印每一行，忽略None行
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0] is not None:
            print(row[0])


def read_excel_to_dicts(
    file_content: io.BytesIO, mapping: dict, header_row: int = 1
) -> list[dict]:
    wb = load_workbook(file_content)
    ws = wb.active

    # 读取标题行
    headers = [mapping.get(cell.value, cell.value) for cell in ws[header_row]]

    data = []
    # 从标题行的下一行开始读取数据
    for row in ws.iter_rows(min_row=header_row + 1):
        if all(cell.value is None for cell in row):
            continue
        row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
        data.append(row_data)
    return data


def insert_data_from_dicts(
    db: Session, model: Type[DeclarativeMeta], data_list: list[dict]
):
    for data in data_list:
        db_item = model(**data)
        db.add(db_item)

    db.commit()
