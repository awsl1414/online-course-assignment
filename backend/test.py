import json
from random import randint

from openpyxl import load_workbook

wb = load_workbook("backend/course_main.xlsx")
work = wb.worksheets[0]
print(work)

ws = wb.active


max_row = ws.max_row

# for j in range(max_row):
#     c_list = ws[j]
#     for i in c_list:
#         print(i.value)

# a = ws.cell(row=1, column=3).value
# print(a)
# print(ws.iter_rows(min_row=1, max_row=2))

for rows in ws.iter_rows(min_row=1, max_row=max_row):
    # print(rows)
    list_a = []
    for row in rows:
        list_a.append(row.value)
        # print(row.cell(column=3).value)
print(list_a)


# a_list = ws[1]
# for i in a_list:
#     print(i.value)
# for j in i:
#     print(j.value)

# for x in range(1, 101):
#     for y in range(1, 101):
#         ws.cell(row=x, column=y, value=10)
# d = ws.cell(row=4, column=2, value=10)
# print(d)
# print(d.value)

# for row in ws.values:
#     for value in row:
#         print(value)

# wb.save("backend/balances.xlsx")
